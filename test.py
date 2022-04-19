from openpyxl import load_workbook
loc =  'static/Output Data Structure.xlsx'
my_wb_obj = load_workbook(filename=loc)
my_sheet_obj = my_wb_obj.active
print(my_sheet_obj.cell(2,2).value)