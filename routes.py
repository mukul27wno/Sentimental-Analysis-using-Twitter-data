from app import app 
from flask import render_template
import tweepy 
from textblob import TextBlob
from openpyxl import load_workbook
from wordcloud import WordCloud
import pandas as pd
import re
import os
import forms

@app.route('/',methods=['GET','POST'])
@app.route('/index.html', methods=['GET','POST'])


def index():
    def calculatee(hashtag,noft,lang,type,i):
        text = ''
        hashtag = hashtag+type
        for q in hashtag:
            for tweets in api.search_tweets(q=q,lang=lang,count=noft):
                text = text + tweets.text
        print("tweet part done")

        loc = 'static/textfiles/'+type+'.txt'
        fp = open(f'{loc}','r')
        text = text + fp.read()
        fp.close()
        print("reading file Done")

        
        text = re.sub('@[A-Za-z0-9]+','',text)  #remove @mentions
        text = re.sub(r'#', '',text)    #remove the '#' symbol
        text = re.sub(r'RT[\s]','',text)    #remove RT
        text = re.sub(r'https?:\/\/\S+','',text)    #remove link
        emoji_pattern = re.compile("["
                           u"\U0001F600-\U0001F64F"  # emoticons
                           u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                           u"\U0001F680-\U0001F6FF"  # transport & map symbols
                           u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                           u"\U00002702-\U000027B0"
                           u"\U000024C2-\U0001F251"
                           "]+", flags=re.UNICODE)
        text = emoji_pattern.sub(r'', string=text)
        print("Cleaning done")

        pl = TextBlob(text).sentiment.polarity
        su = TextBlob(text).sentiment.subjectivity
        if pl <= 0:
            status = 'Negative'
        else:
            status = 'Positive'

        if pl<=0:
            pos = -1-pl
            neg = 1+pl
        else:
            pos = 1-pl
            neg = -1+pl

        sentences = text.split(".")
        sumsen = 0
        tns = 0
        for y in sentences:
            tns = tns + 1  # Total Number of Sentences

        words = text.split(" ")
        tnw = 0
        wordscopy = words

        punctuations = '''!()-[]{};:'"\,<>./?@#$%^&*_~'''

        wordscopy2 = ""
        for char in wordscopy:
            if char not in punctuations:
                wordscopy2 = wordscopy2 + char

        for x in wordscopy:
            tnw = tnw + 1  # Total number of words

        ccw = 0
        cwp = ['un', 'non', 'in', 'pre', 'trans', 're', 'con']
        cws = ['ly', 'ist', 'er', 'ness', 'ment', 's', 'ing', 'ed', 'en', 'est', 'mit', 'ceive', 'fer']
        for y in words:
            for y2 in cwp:
                if y.startswith(y2):
                    ccw = ccw + 1
            for y2 in cws:
                if y.endswith(y2):
                    ccw = ccw + 1

        pcw = ccw / tnw

        fi = 0.4 * ((tnw / tns) + 100 * (ccw / tnw))

        anws = tnw / tns

        sc = 0
        for w in wordscopy2:
            if (
                    w == 'a' or w == 'e' or w == 'i' or w == 'o' or w == 'u' or w == 'A' or w == 'E' or w == 'I' or w == 'O' or w == 'U'):
                sc = sc + 1

        scw = sc / tnw

        pronounRegex = re.compile(
            r'\b(I|me|mine|myself|us|our|ourselves|you|your|yours|yourself|yourselves|he|him|himself|his|she|her|hers|herself|it|its|itself|they|them|their|theirs|themselves|we|my|ours|(?-i:us))\b',
            re.I)
        pronouns = pronounRegex.findall(str(text))

        print("Values Calculated!")

        asl = len(wordscopy2) / tns
        awl = len(wordscopy2) / tnw
        loc =  'static/Output Data Structure.xlsx'
        my_wb_obj = load_workbook(filename=loc)
        my_sheet_obj = my_wb_obj.active
        srocell = my_sheet_obj.cell(row=i,column=1)
        srocell.value = i-1
        typecell = my_sheet_obj.cell(row=i,column=2)
        typecell.value = type
        poscell = my_sheet_obj.cell(row = i, column = 3)
        poscell.value = pos
        negcell = my_sheet_obj.cell(row = i, column = 4)
        negcell.value = neg
        pscell = my_sheet_obj.cell(row = i, column = 5)
        pscell.value = pl
        sucell = my_sheet_obj.cell(row = i, column = 6)
        sucell.value = su
        aslcell = my_sheet_obj.cell(row = i,column = 7)
        aslcell.value = asl
        pcwcell = my_sheet_obj.cell(row = i,column = 8)
        pcwcell.value = pcw
        ficell = my_sheet_obj.cell(row = i,column = 9)
        ficell.value = fi
        anwscell = my_sheet_obj.cell(row = i,column = 10)
        anwscell.value = anws
        cnccell = my_sheet_obj.cell(row = i,column = 11)
        cnccell.value = ccw
        wccell = my_sheet_obj.cell(row = i,column = 12)
        wccell.value = tnw
        sccell = my_sheet_obj.cell(row = i,column = 13)
        sccell.value = scw
        ppcell = my_sheet_obj.cell(row = i,column = 14)
        ppcell.value = len(pronouns)
        awlcell = my_sheet_obj.cell(row = i,column = 15)
        awlcell.value = awl
        my_wb_obj.save(loc)
        print("All values done")

        
        loc = 'static/textfiles/'+type+'.txt'
        fp = open(f'{loc}','w+')  
        fp.write(wordscopy2)
        fp.close()
        wordcloud = WordCloud(max_font_size=50, max_words=100, background_color="white").generate(text)
        loc = 'static/wordcloud/wordcloud-'+type+'.png'
        wordcloud.to_file(loc)
        print("Wordcloud Done")


    log = pd.read_excel("login_twitter_api.xlsx")
    key = log['address']
    ckey = key[0]
    cskey = key[1]
    at = key[2]
    ats = key[3]

    form = forms.AddTaskForm()
    if form.validate_on_submit():
        auth = tweepy.OAuthHandler(ckey, cskey)
        auth.set_access_token(at, ats)
        api = tweepy.API(auth, wait_on_rate_limit = True)
        hashtag = form.hashtag.data
        noft = form.noft.data
        lang = form.lng.data
        print('Submitted ',hashtag)
        calculatee(hashtag,noft,lang,'CryptoAssets',2)
        print('CryptoAssets Done')
        calculatee(hashtag,noft,lang,'DigitalInfrasturcture',3)
        print('DigitalInfrasturcture Done')
        calculatee(hashtag,noft,lang,'DirectTax',4)
        print('DirectTax Done')
        calculatee(hashtag,noft,lang,'Indirecttax',5)
        print('Indirecttax Done')
        calculatee(hashtag,noft,lang,'InfrastructureEnhancement',6)
        print('InfrastructureEnhancement Done')
        calculatee(hashtag,noft,lang,'NationalTeleMedicineProgramme',7)
        print('NationalTeleMedicineProgramme Done')
        calculatee(hashtag,noft,lang,'Sustainability',8)
        print('Sustainability Done')
        loc =  'static/Output Data Structure.xlsx'
        my_wb_obj = load_workbook(filename=loc)
        my_sheet_obj = my_wb_obj.active
        # return render_template('index.html',form=form,hashtag=form.hashtag.data,text=text,Sub=su,Pol=pl,status=status,tns=tns,tnw=tnw,ccw=ccw,pcw=pcw,fi=fi,anws=anws,sc=sc,scw=scw,pronouns=len(pronouns),asl=asl,awl=awl)
        # return render_template('index.html',form=form,hashtag=form.hashtag.data,text=text,Sub=su,Pol=pl,loc=full_filename)
        return render_template('index.html',form=form,hashtag=hashtag,noft=noft,lang=lang,sheet=my_sheet_obj)
    return render_template('index.html',form=form,hashtag=form.hashtag.data)

